import openpyxl


def get_row_count(file,sheet):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.max_row


def get_column_count(file,sheet):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.max_column


def read_data(file,sheet,reading_row,reading_col):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.cell(row=reading_row,column=reading_col).value


def write_data(file,sheet,reading_row,reading_col,data):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    worksheet.cell(row=reading_row,column=reading_col).value = data
    workbook.save(file)

